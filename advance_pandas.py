import signal
import pandas as pd
from os import startfile
from pathlib import Path
from typing import Optional, Union
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile
from shutil import move, copy2
from multiprocessing import Process
from tkinter import filedialog, messagebox
from .utilities import wait_for_file_availability, always_on_top_dialog


class AdvancePandas(pd.DataFrame):
    """
    A subclass of pandas DataFrame that adds advanced save functionality,
    including async saving, format retention, and automatic backups.
    """
    _metadata = ["source_file", "destination_file"]
    
    def __init__(self, dataframe=None, source_file: Optional[Union[str, Path]] = None, destination_file: Optional[Union[str, Path]] = None, **kwargs):
        """
        Initialize the AdvancePandas object with optional source and destination file paths.
        
        Parameters:
        - dataframe: The initial DataFrame content.
        - source_file: Path to the original file (if any).
        - destination_file: Path to where the file should be saved.
        """
        self.source_file = source_file
        self.destination_file = destination_file
        super().__init__(dataframe, **kwargs)
    
    def save(self, file_path: Optional[Union[str, Path]] = None, async_mode: bool = False, retain_format: bool = False, auto_open: bool = False, create_backup: bool = False):
        """
        Save the DataFrame to a file with optional formatting, async mode, and backup creation.
        
        Parameters:
        - file_path: The target file path for saving. Defaults to source or destination file if not provided.
        - async_mode: If True, the save operation will run in a separate process.
        - retain_format: If True, preserves column widths and alignments for Excel files.
        - auto_open: If True, opens the saved file after completion.
        - create_backup: If True, creates a backup copy of the saved file.
        """
        if not file_path:
            if self.destination_file:
                file_path = self.destination_file
            else:
                if self.source_file:
                    if always_on_top_dialog(messagebox.askyesno, "AdvancePandas Notice", f"No destination file found.\nSave to source file?\nSource File: {self.source_file}"):
                        file_path = self.source_file
                    else:
                        filetypes = [
                            ("Excel and CSV files", "*.xlsx;*.csv"),
                            ("Excel files", "*.xlsx"),
                            ("CSV files", "*.csv"),
                            ("All files", "*.*")
                        ]
                        file_path = always_on_top_dialog(
                            filedialog.asksaveasfilename,
                            defaultextension=".xlsx",
                            filetypes=filetypes,
                            title="Save File As",
                        )
                if not file_path:
                    raise ValueError("No save file_path provided.")
        
        if async_mode:
            process = Process(target=self._save_to_file, args=(file_path,), kwargs={"retain_format": retain_format, "auto_open": auto_open, "create_backup": create_backup})
            process.start()
        else:
            self._save_to_file(file_path, retain_format, auto_open, create_backup)

    def _save_to_file(self, file_path: Union[str, Path], retain_format: bool = False, auto_open: bool = False, create_backup: bool = False):
        """
        Internal method for saving the DataFrame while ensuring format retention and async safety.
        
        Parameters:
        - file_path: The destination path for saving the file.
        - retain_format: Whether to preserve formatting for Excel files.
        - auto_open: Whether to open the file after saving.
        - create_backup: Whether to create a backup of the saved file.
        """
        file_path = Path(file_path)
        
        def ignore_keyboard_interrupt():
            pass
        
        signal.signal(signal.SIGINT, ignore_keyboard_interrupt)
        valid_extensions = ['.csv', '.xlsx']
        if file_path.suffix not in valid_extensions:
            raise ValueError(f"File extension must be one of {', '.join(valid_extensions)}.")
        
        try:
            temp_file_dir = file_path.parent / 'TemporaryFiles'
            temp_file_dir.mkdir(exist_ok=True)
            
            with NamedTemporaryFile(delete=False, suffix=file_path.suffix, dir=temp_file_dir) as temp_file:
                temp_file_path = Path(temp_file.name)
                
                if file_path.suffix == '.csv':
                    self.to_csv(temp_file_path, index=False)
                elif file_path.suffix == '.xlsx':
                    self.to_excel(temp_file_path, index=False)
                
            if retain_format:
                if reference_file := file_path if file_path.exists() else self.source_file:
                    if Path(reference_file).suffix in ['.xlsx', '.xls']:
                        AdvancePandas._transfer_excel_format(reference_file, temp_file_path, self)
                
            if file_path.exists():
                wait_for_file_availability(file_path)
            move(temp_file_path, file_path)
            
            if create_backup:
                backup_file_path = file_path.with_name(f"{file_path.stem} - Backup{file_path.suffix}")
                if backup_file_path.exists() and wait_for_file_availability(backup_file_path, notify_only=True):
                    copy2(file_path, backup_file_path)
            
            if auto_open:
                startfile(file_path)
        
        finally:
            signal.signal(signal.SIGINT, signal.default_int_handler)
    
    @staticmethod
    def _transfer_excel_format(source_file: Path, temp_file_path: Path, dataframe: pd.DataFrame):
        """Extracts and applies formatting from the source Excel file to the new file."""
        workbook = load_workbook(source_file)
        sheet = workbook.active

        # Extract column widths and alignments
        column_details = {
            get_column_letter(col_idx): {
                "width": sheet.column_dimensions[get_column_letter(col_idx)].width,
                "alignment": sheet.cell(row=1, column=col_idx).alignment
            }
            for col_idx in range(1, sheet.max_column + 1)
        }

        with pd.ExcelWriter(temp_file_path, engine="openpyxl") as writer:
            dataframe.to_excel(writer, sheet_name=sheet.title, index=False, header=True)

        updated_workbook = load_workbook(temp_file_path)
        updated_sheet = updated_workbook[sheet.title]

        for col_idx, col_letter in enumerate(column_details.keys(), start=1):
            updated_sheet.column_dimensions[col_letter].width = column_details[col_letter]["width"]
            for cell in updated_sheet[get_column_letter(col_idx)]:
                cell.alignment = column_details[col_letter]["alignment"]

        updated_workbook.save(temp_file_path)

    @property
    def _constructor(self):
        return AdvancePandas
    

def AdvanceExcelReader(source_file: Union[str, Path], destination_file: Optional[Union[str, Path]] = None, continue_from_saved: bool = False) -> AdvancePandas:
    """
    Read an Excel or CSV file into an AdvancePandas instance.
    
    Parameters:
    - source_file: The path to the source Excel or CSV file.
    - destination_file: The path to save modifications. If continue_from_saved is True, this may override source_file.
    - continue_from_saved: If True, attempts to continue from the saved destination file if available.
    """
    source_file = Path(source_file)
    if not source_file.exists() and not continue_from_saved:
        raise FileNotFoundError(f"The file {source_file} does not exist.")
    
    if continue_from_saved and destination_file:
        destination_file = Path(destination_file)
        if destination_file.exists():
            source_file = destination_file
    
    df = pd.read_csv(source_file) if source_file.suffix == ".csv" else pd.read_excel(source_file)
    
    return AdvancePandas(df, source_file=source_file, destination_file=destination_file)
